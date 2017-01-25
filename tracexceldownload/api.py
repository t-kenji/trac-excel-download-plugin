# -*- coding: utf-8 -*-

import inspect
from cStringIO import StringIO
from datetime import datetime
from decimal import Decimal
from unicodedata import east_asian_width
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import xlwt
except ImportError:
    xlwt = None

from trac.core import Component, TracError
from trac.util.text import to_unicode
from tracexceldownload.translation import ChoiceOption, N_, ngettext


__all__ = ('get_excel_format', 'get_excel_mimetype', 'get_workbook_writer')


def get_excel_format(env):
    format = ExcelDownloadConfig(env).format
    if format == '(auto)':
        if openpyxl:
            return 'xlsx'
        if xlwt:
            return 'xls'
        raise TracError("Require openpyxl or xlwt library")
    if format == 'xlsx':
        if openpyxl:
            return format
        raise TracError("Require openpyxl library")
    if format == 'xls':
        if xlwt:
            return format
        raise TracError("Require xlwt library")
    raise TracError("Unsupported format: '%s'" % format)


def _writer(ext):
    for cls in (OpenpyxlWorkbookWriter, XlwtWorkbookWriter):
        if cls.ext == ext:
            return cls
    raise TracError("Unsupported format '%s'" % ext)


def get_excel_mimetype(ext):
    return _writer(ext).mimetype


def get_workbook_writer(env, req):
    ext = get_excel_format(env)
    cls = _writer(ext)
    return cls(env, req)


class ExcelDownloadConfig(Component):

    format = ChoiceOption('exceldownload', 'format', ('(auto)', 'xlsx', 'xls'),
        doc=N_("Specifies the format of Excel file to download."))


class WorksheetWriterError(TracError): pass


class AbstractWorkbookWriter(object):

    ext = None
    mimetype = None

    def __init__(self, env, req, book):
        self.env = env
        self.log = env.log
        self.req = req
        self.tz = req.tz
        if hasattr(req, 'locale'):
            self.ambiwidth = (1, 2)[str(req.locale)[:2] in ('ja', 'kr', 'zh')]
        else:
            self.ambiwidth = 1
        self.book = book
        self.styles = self._get_excel_styles()
        self._metrics_cache = {}

    def create_sheet(self, title):
        raise NotImplemented

    def dump(self, out):
        raise NotImplemented

    def dumps(self):
        out = StringIO()
        self.dump(out)
        return out.getvalue()

    def _get_excel_styles(self):
        raise NotImplemented

    def get_metrics(self, value):
        if not value:
            return 0, 1
        if isinstance(value, str):
            value = to_unicode(value)
        if value not in self._metrics_cache:
            lines = value.splitlines()
            doubles = ('WFA', 'WF')[self.ambiwidth == 1]
            width = max(sum((1, 2)[east_asian_width(ch) in doubles]
                            for ch in line)
                        for line in lines)
            if len(value) > 64:
                return width, len(lines)
            self._metrics_cache[value] = (width, len(lines))
        return self._metrics_cache[value]


class AbstractWorksheetWriter(object):

    MAX_ROWS = None
    MAX_COLS = None
    MAX_CHARS = None

    def __init__(self, sheet, writer):
        self.writer = writer
        self.book = writer.book
        self.sheet = sheet
        self.styles = self.writer.styles
        self.row_idx = 0
        self._col_widths = {}
        self.tz = writer.req.tz

    def write_row(self, cells):
        raise NotImplemented

    def move_row(self):
        self.row_idx += 1
        if self.row_idx >= self.MAX_ROWS:
            raise max_rows_error(self.MAX_ROWS)

    def get_metrics(self, value):
        return self.writer.get_metrics(value)

    def _set_col_width(self, idx, width):
        widths = self._col_widths
        widths.setdefault(idx, 1)
        if widths[idx] < width:
            widths[idx] = width

    def set_col_widths(self):
        raise NotImplemented

    def _normalize_text(self, value):
        if isinstance(value, str):
            value = to_unicode(value)
        value = '\n'.join(line.rstrip() for line in value.splitlines())
        if len(value) > self.MAX_CHARS:
            value = value[:self.MAX_CHARS - 1] + u'\u2026'
        return value


class OpenpyxlWorkbookWriter(AbstractWorkbookWriter):

    ext = 'xlsx'
    mimetype = 'application/' \
               'vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def __init__(self, env, req):
        if not openpyxl:
            raise TracError('Require openpyxl library')
        book = self._create_book()
        AbstractWorkbookWriter.__init__(self, env, req, book)
        for style in self.styles.itervalues():
            book.add_named_style(style)

    def create_sheet(self, title):
        sheet = self.book.create_sheet(title=title)
        return OpenpyxlWorksheetWriter(sheet, self)

    def dump(self, out):
        self.book.save(out)

    def _get_excel_styles(self):
        from openpyxl.styles import (
            Alignment, Border, Font, NamedStyle, PatternFill, Side)

        def create_font(**kwargs):
            if 'size' not in kwargs:
                kwargs['size'] = 9
            return Font(name='Arial', **kwargs)

        def style_base(name):
            style = NamedStyle(name=name)
            style.font = create_font(size=9)
            style.alignment = Alignment(vertical='top', wrap_text=True)
            side = Side(style='thin')
            style.border = Border(left=side, right=side, top=side, bottom=side)
            return style

        def style_change(fn):
            style = fn()
            style.name = '%s:change' % style.name
            style.fill = PatternFill(fill_type='solid', fgColor='00FF9900')
            return style

        def style_header():
            style = NamedStyle(name='header')
            style.font = create_font(size=20)
            return style

        def style_header2():
            style = NamedStyle(name='header2')
            style.font = create_font(size=16)
            return style

        def style_thead():
            style = NamedStyle(name='thead')
            style.font = create_font(bold=True, color='00FFFFFF')
            style.fill = PatternFill(patternType='solid', fgColor='00000000')
            side = Side(style='thin', color='00FFFFFF')
            style.border = Border(left=side, right=side, top=side, bottom=side)
            return style

        def style_id():
            style = style_base('id')
            style.alignment = Alignment(vertical='top', horizontal='right')
            style.number_format = '"#"0'
            return style

        def style_milestone():
            style = style_base('milestone')
            style.number_format = '@'
            return style

        def style_time():
            style = style_base('[time]')
            style.number_format = 'HH:MM:SS'
            return style

        def style_date():
            style = style_base('[date]')
            style.number_format = 'YYYY-MM-DD'
            return style

        def style_datetime():
            style = style_base('[datetime]')
            style.number_format = 'YYYY-MM-DD HH:MM:SS'
            return style

        def style_default():
            style = style_base('*')
            style.number_format = '@'
            return style

        styles = []
        for fn in (style_header, style_header2, style_thead,
                   style_id, style_milestone, style_time, style_date,
                   style_datetime, style_default):
            styles.append(fn())
            styles.append(style_change(fn))
        return dict((style.name, style) for style in styles)

    if not openpyxl:
        def _create_book(self):
            raise NotImplemented
    elif 'write_only' in inspect.getargspec(openpyxl.Workbook.__init__)[0]:
        def _create_book(self):
            return openpyxl.Workbook(write_only=True)
    else:
        def _create_book(self):
            return openpyxl.Workbook(optimized_write=True)


class OpenpyxlWorksheetWriter(AbstractWorksheetWriter):

    MAX_ROWS = 1048576
    MAX_COLS = 16384
    MAX_CHARS = 32767

    def __init__(self, sheet, writer):
        AbstractWorksheetWriter.__init__(self, sheet, writer)
        self._rows = []

    def write_row(self, cells):
        get_metrics = self.get_metrics
        tz = self.tz
        has_tz_normalize = hasattr(tz, 'normalize')  # pytz

        values = []
        for idx, (value, style, width, line) in enumerate(cells):
            if isinstance(value, datetime):
                value = value.astimezone(tz)
                if has_tz_normalize:
                    value = tz.normalize(value)
                value = datetime(*(value.timetuple()[0:6]))
                if style == '[date]':
                    width = len('YYYY-MM-DD')
                elif style == '[time]':
                    width = len('HH:MM:SS')
                else:
                    width = len('YYYY-MM-DD HH:MM:SS')
                width /= 1.2
                line = 1
            elif isinstance(value, (int, long, float, Decimal)):
                width = len('%g' % value) / 1.2
                line = 1
            elif value is True or value is False:
                width = 5 / 1.2
                line = 1
            elif isinstance(value, basestring):
                value = self._normalize_text(value)

            if width is None or line is None:
                metrics = get_metrics(value)
                if width is None:
                    width = metrics[0]
                if line is None:
                    line = metrics[1]

            cell = OpenpyxlCell(value)
            if style not in self.styles:
                if style.endswith(':change'):
                    style = '*:change'
                else:
                    style = '*'
            cell.style = style
            values.append(cell)
            self._set_col_width(idx, width)

        self._rows.append(values or (None,))
        self.row_idx += 1

    def set_col_widths(self):
        from openpyxl.utils.cell import get_column_letter
        from openpyxl.cell import Cell

        for idx, width in sorted(self._col_widths.iteritems()):
            letter = get_column_letter(idx + 1)
            self.sheet.column_dimensions[letter].width = 1 + min(width, 50)
        for row in self._rows:
            values = []
            for val in row:
                if val:
                    cell = Cell(self.sheet, column='A', row=1, value=val.value)
                    cell.style = val.style
                else:
                    cell = val
                values.append(cell)
            self.sheet.append(values)


class OpenpyxlCell(object):

    __slots__ = ('value', 'style')

    def __init__(self, value, style='*'):
        self.value = value
        self.style = style


class XlwtWorkbookWriter(AbstractWorkbookWriter):

    ext = 'xls'
    mimetype = 'application/vnd.ms-excel'

    def __init__(self, env, req):
        if not xlwt:
            raise TracError('Require xlwt library')
        book = xlwt.Workbook(encoding='utf-8', style_compression=1)
        AbstractWorkbookWriter.__init__(self, env, req, book)

    def create_sheet(self, title):
        sheet = self.book.add_sheet(title)
        return XlwtWorksheetWriter(sheet, self)

    def dump(self, out):
        self.book.save(out)

    def _get_excel_styles(self):
        Alignment = xlwt.Alignment
        SOLID_PATTERN = xlwt.Pattern.SOLID_PATTERN
        THIN = xlwt.Borders.THIN
        UNDERLINE_SINGLE = xlwt.Font.UNDERLINE_SINGLE
        XFStyle = xlwt.XFStyle
        colour_map = xlwt.Style.colour_map

        def style_base():
            style = XFStyle()
            style.alignment.vert = Alignment.VERT_TOP
            style.alignment.wrap = True
            style.font.height = 180 # 9pt
            borders = style.borders
            borders.left = THIN
            borders.right = THIN
            borders.top = THIN
            borders.bottom = THIN
            return style

        header = XFStyle()
        header.font.height = 400 # 20pt
        header2 = XFStyle()
        header2.font.height = 320 # 16pt

        thead = style_base()
        thead.font.bold = True
        thead.font.colour_index = colour_map['white']
        thead.pattern.pattern = SOLID_PATTERN
        thead.pattern.pattern_fore_colour = colour_map['black']
        thead.borders.colour = 'white'
        thead.borders.left = THIN
        thead.borders.right = THIN
        thead.borders.top = THIN
        thead.borders.bottom = THIN

        def style_change(style):
            pattern = style.pattern
            pattern.pattern = SOLID_PATTERN
            pattern.pattern_fore_colour = colour_map['light_orange']
            return style

        def style_id():
            style = style_base()
            style.font.underline = UNDERLINE_SINGLE
            style.font.colour_index = colour_map['blue']
            style.num_format_str = '"#"0'
            return style

        def style_milestone():
            style = style_base()
            style.font.underline = UNDERLINE_SINGLE
            style.font.colour_index = colour_map['blue']
            style.num_format_str = '@'
            return style

        def style_time():
            style = style_base()
            style.num_format_str = 'HH:MM:SS'
            return style

        def style_date():
            style = style_base()
            style.num_format_str = 'YYYY-MM-DD'
            return style

        def style_datetime():
            style = style_base()
            style.num_format_str = 'YYYY-MM-DD HH:MM:SS'
            return style

        def style_default():
            style = style_base()
            style.num_format_str = '@'    # String
            return style

        styles = {'header': header, 'header2': header2, 'thead': thead}
        for key, func in (('id', style_id),
                          ('milestone', style_milestone),
                          ('[time]', style_time),
                          ('[date]', style_date),
                          ('[datetime]', style_datetime),
                          ('*', style_default)):
            styles[key] = func()
            styles['%s:change' % key] = style_change(func())
        return styles


class XlwtWorksheetWriter(AbstractWorksheetWriter):

    MAX_ROWS = 65536
    MAX_COLS = 255
    MAX_CHARS = 32767

    def __init__(self, sheet, writer):
        AbstractWorksheetWriter.__init__(self, sheet, writer)
        self._cells_count = 0

    def move_row(self):
        AbstractWorksheetWriter.move_row(self)
        self._flush_row()

    def write_row(self, cells):
        _get_style = self._get_style
        _set_col_width = self._set_col_width
        get_metrics = self.get_metrics
        tz = self.tz
        has_tz_normalize = hasattr(tz, 'normalize')  # pytz

        row = self.sheet.row(self.row_idx)
        max_line = 1
        max_height = 0
        for idx, (value, style, width, line) in enumerate(cells):
            if isinstance(value, datetime):
                value = value.astimezone(tz)
                if has_tz_normalize:
                    value = tz.normalize(value)
                value = datetime(*(value.timetuple()[0:6]))
                if style == '[date]':
                    width = len('YYYY-MM-DD')
                elif style == '[time]':
                    width = len('HH:MM:SS')
                else:
                    width = len('YYYY-MM-DD HH:MM:SS')
                _set_col_width(idx, width)
                row.set_cell_date(idx, value, _get_style(style))
                continue
            if isinstance(value, (int, long, float, Decimal)):
                _set_col_width(idx, len('%g' % value))
                row.set_cell_number(idx, value, _get_style(style))
                continue
            if value is True or value is False:
                _set_col_width(idx, 1)
                row.set_cell_number(idx, int(value), _get_style(style))
                continue
            if isinstance(value, basestring):
                value = self._normalize_text(value)
            if width is None or line is None:
                metrics = get_metrics(value)
                if width is None:
                    width = metrics[0]
                if line is None:
                    line = metrics[1]
            if max_line < line:
                max_line = line
            _set_col_width(idx, width)
            style = _get_style(style)
            if max_height < style.font.height:
                max_height = style.font.height
            row.write(idx, value, style)
            self._cells_count += 1
        row.height = min(max_line, 10) * max(max_height * 255 / 180, 255)
        row.height_mismatch = True
        self.move_row()

    def _flush_row(self):
        if self.row_idx % 512 == 0 or self._cells_count >= 4096:
            self.sheet.flush_row_data()
            self._cells_count = 0

    def _get_style(self, style):
        if isinstance(style, basestring):
            if style not in self.styles:
                if style.endswith(':change'):
                    style = '*:change'
                else:
                    style = '*'
            style = self.styles[style]
        return style

    def set_col_widths(self):
        for idx, width in self._col_widths.iteritems():
            self.sheet.col(idx).width = (1 + min(width, 50)) * 256


def max_rows_error(num):
    message = ngettext(
        "Number of rows in the Excel sheet exceeded the limit of %(num)d row",
        "Number of rows in the Excel sheet exceeded the limit of %(num)d rows",
        num)
    return WorksheetWriterError(message)
